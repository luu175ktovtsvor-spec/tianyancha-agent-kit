"""Delimited, JSONL, and XLSX exports without provider-specific columns."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import CompanyRecord
from .privacy import redact_record

BASE_COLUMNS = (
    "name", "credit_code", "entity_type", "established_on", "status", "province", "city", "district",
    "address", "representative", "phones", "business_scope", "categories", "risk", "source_id",
)
PHONE_LEAD_BASE_COLUMNS = (
    "entity_class", "name", "credit_code", "entity_type", "established_on", "status", "province", "city",
    "district", "address", "representative", "phone_count", "business_scope", "categories", "risk", "source_id",
)


def _rows(records: list[CompanyRecord], redactions: list[str]) -> list[dict[str, Any]]:
    rows = []
    for record in records:
        row = redact_record(record, redactions)
        rows.append({column: row.get(column, "") for column in BASE_COLUMNS})
    return rows


def _cell(value: Any) -> Any:
    if isinstance(value, list):
        value = "; ".join(str(item) for item in value)
    elif isinstance(value, dict):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str) and value.lstrip(" \t\r\n").startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def export_records(records: list[CompanyRecord], destination: str | Path, fmt: str, redactions: list[str] | None = None) -> Path:
    """Export canonical records. ``fmt`` is one of ``jsonl``, ``csv``, or ``xlsx``."""

    target = Path(destination)
    target.parent.mkdir(parents=True, exist_ok=True)
    rows = _rows(records, redactions or [])
    if fmt == "jsonl":
        with target.open("w", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    elif fmt == "csv":
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=BASE_COLUMNS)
            writer.writeheader()
            for row in rows:
                writer.writerow({key: _cell(value) for key, value in row.items()})
    elif fmt == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "companies"
        sheet.append(list(BASE_COLUMNS))
        for row in rows:
            sheet.append([_cell(row[column]) for column in BASE_COLUMNS])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 48)
        workbook.save(target)
    else:
        raise ValueError("format must be jsonl, csv, or xlsx")
    return target


def _write_tabular(rows: list[dict[str, Any]], headers: list[str], target: Path, fmt: str, sheet_name: str) -> None:
    if fmt == "jsonl":
        with target.open("w", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
        return
    if fmt == "csv":
        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({key: _cell(row.get(key, "")) for key in headers})
        return
    if fmt == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name[:31]
        sheet.append(headers)
        for row in rows:
            sheet.append([_cell(row.get(header, "")) for header in headers])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 48)
        workbook.save(target)
        return
    raise ValueError("format must be jsonl, csv, or xlsx")


def _phone_lead_rows(
    records: list[tuple[str, CompanyRecord]], redactions: list[str]
) -> tuple[list[dict[str, Any]], list[str]]:
    rendered: list[tuple[str, dict[str, Any]]] = []
    max_phones = 0
    redact_phones = "phones" in redactions
    for entity_class, record in records:
        values = redact_record(record, redactions)
        phones = values.get("phones") if isinstance(values.get("phones"), list) else []
        max_phones = max(max_phones, len(phones))
        rendered.append((entity_class, values))
    phone_columns = [f"phone_{index}" for index in range(1, max(max_phones, 1) + 1)]
    headers = [*PHONE_LEAD_BASE_COLUMNS[:11], *phone_columns, *PHONE_LEAD_BASE_COLUMNS[11:]]
    rows: list[dict[str, Any]] = []
    for entity_class, values in rendered:
        phones = values.get("phones") if isinstance(values.get("phones"), list) else []
        row = {key: values.get(key, "") for key in PHONE_LEAD_BASE_COLUMNS if key not in {"entity_class", "phone_count"}}
        row["entity_class"] = entity_class
        row["phone_count"] = "[REDACTED]" if redact_phones else len(phones)
        row.update({column: phones[index] if index < len(phones) else "" for index, column in enumerate(phone_columns)})
        rows.append(row)
    return rows, headers


def export_phone_leads(
    leads_by_entity: dict[str, list[CompanyRecord]],
    output_dir: str | Path,
    fmt: str,
    redactions: list[str] | None = None,
) -> dict[str, Path]:
    """Export acquisition leads into an all-leads file and one file per entity class."""

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    extension = {"jsonl": "jsonl", "csv": "csv", "xlsx": "xlsx"}.get(fmt)
    if not extension:
        raise ValueError("format must be jsonl, csv, or xlsx")
    selected_redactions = redactions or []
    outputs: dict[str, Path] = {}
    all_records = [(entity_class, record) for entity_class, records in leads_by_entity.items() for record in records]
    all_rows, all_headers = _phone_lead_rows(all_records, selected_redactions)
    all_path = destination / f"all_phone_leads.{extension}"
    _write_tabular(all_rows, all_headers, all_path, fmt, "all_phone_leads")
    outputs["all"] = all_path
    for entity_class, records in leads_by_entity.items():
        rows, headers = _phone_lead_rows([(entity_class, record) for record in records], selected_redactions)
        path = destination / f"{entity_class}_phone_leads.{extension}"
        _write_tabular(rows, headers, path, fmt, entity_class)
        outputs[entity_class] = path
    return outputs
